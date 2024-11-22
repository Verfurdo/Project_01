# data_analysis.py - Adatbetöltés és előkészítés / Statisztikai elemzések számítása és mentése

import pandas as pd  # Adatkezeléshez, DataFrame-ek létrehozásához
from scipy import stats  # Statisztikai számításokhoz, (korrelációs együttható)
from sklearn.linear_model import LinearRegression  # Lineáris regressziós modell
from sklearn.metrics import mean_squared_error, mean_absolute_error  # Hibatípusok kiszámításához
import numpy as np  # Numerikus számítások és tömbök kezelése
import os  # Operációs rendszerrel kapcsolatos műveletekhez, (elérési út kezelése)
import streamlit as st  # Streamlit modulok importálása
import time  # Az idő kezelésére szolgáló modul importálása (késleltetés)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

# Adatok betöltése és előkészítése
def adatok_betoltese_elokeszitese():
    # Az analysis.py fájl könyvtára alapján az adatfájl csv meghatározása
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  
    data_path = os.path.join(script_dir, 'data', 'stadat-nep0013-22.1.1.13-hu.csv')  # Adatfájl elérési útja

    # Fájl létezésének ellenőrzése Streamlit figyelmeztetéssel
    if not os.path.exists(data_path):
        st.error(f"Az elemzéshez szükséges CSV fájl nem található 'stadat-nep0013-22.1.1.13-hu.csv' Helyezd a fájlt a 'data' mappába, és próbáld újra.")
        st.stop()  # A Streamlit futásának megszakítása, ha a fájl hiányzik

    # CSV fájl betöltése DataFrame-be, elválasztó karakter és karakterkódolás megadása
    df = pd.read_csv(data_path, sep=";", encoding="ISO-8859-1")  

    # Tisztítás és átalakítás: a felesleges szóközök és vesszők cseréje
    df = df.replace(" ", "", regex=True)  # Felesleges szóközök eltávolítása
    df = df.replace(',', '.', regex=True)  # Vesszők pontokra cserélése, hogy float típusú lehessen
    df['Termelés'] = pd.to_numeric(df['Termelés'], errors='coerce')  # 'Termelés' oszlop átalakítása numerikussá
    df['Fogyasztás'] = pd.to_numeric(df['Fogyasztás'], errors='coerce')  # 'Fogyasztás' oszlop átalakítása numerikussá
    df['Év'] = pd.to_numeric(df['Év'], errors='coerce')  # 'Év' oszlop átalakítása numerikussá

    return df  # Előkészített DataFrame visszaadása

# Statisztikai elemzések számítása és mentése a 'data' mappába
def statisztikai_elemzes_mentese(df):
    # Eredmények mentésére szolgáló elérési út meghatározása
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_path = os.path.join(script_dir, 'data', 'statistical_analysis.xls') # Mentési fájl helye

    rows = []  # Lista az összes statisztikai adat tárolására

    # Általános statisztikai mutatók számítása és tárolása
    for column in df.columns:
        if pd.api.types.is_numeric_dtype(df[column]):
            rows.append({'Oszlop': column, 'Metrika': 'Átlag', 'Érték': np.mean(df[column])})
            rows.append({'Oszlop': column, 'Metrika': 'Medián', 'Érték': np.median(df[column])})
            rows.append({'Oszlop': column, 'Metrika': 'Szórás', 'Érték': np.std(df[column])})
            rows.append({'Oszlop': column, 'Metrika': 'Variancia', 'Érték': np.var(df[column])})
            rows.append({'Oszlop': column, 'Metrika': 'Minimum', 'Érték': np.min(df[column])})
            rows.append({'Oszlop': column, 'Metrika': 'Maximum', 'Érték': np.max(df[column])})

    # Korrelációs együttható és p-érték számítása a 'Termelés' és 'Fogyasztás' oszlopok között
    X = df['Termelés'].values.reshape(-1, 1)  # 'Termelés' oszlop átalakítása a regresszióhoz
    y = df['Fogyasztás'].values  # 'Fogyasztás' oszlop célértékként

    # Pearson korrelációs együttható és p-érték kiszámítása
    correlation, p_value = stats.pearsonr(df['Termelés'], df['Fogyasztás'])
    rows.append({'Oszlop': 'Termelés-Fogyasztás', 'Metrika': 'Korrelációs együttható', 'Érték': correlation})
    rows.append({'Oszlop': 'Termelés-Fogyasztás', 'Metrika': 'p-érték', 'Érték': p_value})

    # Lineáris regresszió modell létrehozása és illesztése
    model = LinearRegression()
    model.fit(X, y)  # Modell betanítása az X és y adatokra
    
    # Modell paramétereinek (meredekség és tengelymetszet) kinyerése és tárolása
    slope = model.coef_[0]  # Regressziós egyenes meredeksége
    intercept = model.intercept_  # Y tengelymetszet
    rows.append({'Oszlop': 'Regresszió', 'Metrika': 'Meredekség (slope)', 'Érték': slope})
    rows.append({'Oszlop': 'Regresszió', 'Metrika': 'Y tengelymetszet (intercept)', 'Érték': intercept})

    # Regressziós modell R² értékének kiszámítása és tárolása
    r_squared = model.score(X, y)
    rows.append({'Oszlop': 'Regresszió', 'Metrika': 'R² érték', 'Érték': r_squared})

    # Modell előrejelzései és hibamértékek számítása
    y_pred = model.predict(X)  # Előrejelzések az X alapján
    mae = mean_absolute_error(y, y_pred)  # Mean Absolute Error kiszámítása
    mse = mean_squared_error(y, y_pred)  # Mean Squared Error kiszámítása
    rows.append({'Oszlop': 'Regresszió', 'Metrika': 'MAE (Mean Absolute Error)', 'Érték': mae})
    rows.append({'Oszlop': 'Regresszió', 'Metrika': 'MSE (Mean Squared Error)', 'Érték': mse})

    # Jövőbeli érték előrejelzése a modell alapján
    future_production = np.array([[565]])  # Bemeneti érték a jövőbeli termelésre
    future_consumption_pred = model.predict(future_production)  # Jövőbeli fogyasztás előrejelzése
    rows.append({'Oszlop': 'Előrejelzés', 'Metrika': 'Fogyasztás 2025-re (termelés=565)', 'Érték': future_consumption_pred[0]})

    # Excel fájl mentése
    wb = Workbook()
    ws = wb.active
    ws.title = "Statisztikai Elemzések"
    stats_df = pd.DataFrame(rows)

    # Adatok írása az Excel fájlba
    for row in stats_df.itertuples(index=False, name=None):
        ws.append(row)

    # Formázás hozzáadása
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Mentési próbálkozás és Streamlit üzenetek
    while True:
        try:
            wb.save(output_path)
            
            # Üzenet a fájl mentéséről
            message = st.empty()  
            message.success(f"Statisztikai elemzések számítása mentve a következő fájlba 'data' mappa 'statistical_analysis.xls'")
            
            # Üzenet a diagram készítéséről
            message2 = st.empty()
            message2.info("Vonal- Pontdiagram és statisztikai táblázat készítése a beolvasott adatok alapján...")
            
            time.sleep(7)  # 7 másodperc várakozás
            
            # Üzenetek eltávolítása
            message.empty()
            message2.empty()
            
            break

        except PermissionError:
            st.error(f"HIBA: Nem lehet menteni a fájlt, mert az meg van nyitva: {output_path}")
            st.info("Kérlek, zárd be a fájlt, és próbáld újra a frissítéssel.")
            st.stop()
